"""
Sync US Warehouse catalog prices from Price master US warehouse.xlsx (PRICE_COMPARE).
Updates products.json and index.html CARD_INNER_HTML size-pill data-price values.
Single-vial prices are matched by catalog code (vial image basename).
"""
from __future__ import annotations

import json
import os
import re
import shutil
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
INDEX = ROOT / "index.html"
PRODUCTS = ROOT / "products.json"
XLSX = Path(r"C:\Users\Jason\OneDrive\Desktop\Projects\Pep Vendors\Price master US warehouse.xlsx")
XLSX_FALLBACK = Path(os.environ.get("TEMP", ".")) / "price-master-copy.xlsx"


def fmt_price(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        if float(value).is_integer():
            return f"${int(value)}"
        return f"${value:.2f}"
    s = str(value).strip()
    if s and not s.startswith("$"):
        try:
            n = float(s)
            if n.is_integer():
                return f"${int(n)}"
            return f"${n:.2f}"
        except ValueError:
            pass
    return s


def load_price_compare():
    source = XLSX
    try:
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    except PermissionError:
        shutil.copy2(XLSX, XLSX_FALLBACK)
        source = XLSX_FALLBACK
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    ws = wb["PRICE_COMPARE"]
    rows = list(ws.iter_rows(min_row=3, max_col=5, values_only=True))
    wb.close()
    by_code: dict[str, dict] = {}
    for row in rows:
        code, name, single, kit, spec = (row + (None, None, None, None, None))[:5]
        if not code or not str(code).strip():
            continue
        code = str(code).strip()
        by_code[code] = {
            "name": str(name or "").strip(),
            "single": fmt_price(single),
            "kit": fmt_price(kit),
            "spec": str(spec or "").strip(),
        }
    return by_code


# Research kit cards in index.html: h3 title -> PRICE_COMPARE catalog code(s)
KIT_CARD_CODES: dict[str, str | list[str]] = {
    "BPC-157 10mg Kit": "BC10",
    "BPC-157 + TB-500 Blend 10mg Kit": "BB10",
    "CJC-1295 10mg Kit": "CND10",
    "CJC-1295 DAC 5mg Kit": "CD5",
    "CJC-1295 + Ipamorelin Blend 10mg Kit": "CP10",
    "DSIP 10mg Kit": "DS10",
    "GHK-CU Kit": ["CU50", "CU100"],
    "GLP 1 Kit": ["SM5", "RT5", "TR10", "CGL5"],
    "KLOW 80mg Kit": "KLOW80",
    "KPV 10mg Kit": "KP10",
    "LL-37 5mg Kit": "LL37",
    "MOTS-C 10mg Kit": "MS10",
    "NAD+ 500mg Kit": "NJ500",
    "PT-141 10mg Kit": "P41",
    "TB-500 10mg Kit": "BT10",
    "Tesamorelin 10mg Kit": "TSM10",
}


def parse_index_embeds(html: str):
    card_prefix = "  var CARD_INNER_HTML = "
    card_start = html.index(card_prefix)
    size_markers = [";\r\n  window._sizeImages = ", ";\n  window._sizeImages = "]
    size_start = -1
    size_prefix = ""
    for marker in size_markers:
        pos = html.find(marker, card_start)
        if pos >= 0:
            size_start = pos
            size_prefix = marker
            break
    if size_start < 0:
        raise ValueError("Could not find window._sizeImages marker")
    cards = json.loads(html[card_start + len(card_prefix) : size_start])
    size_json_start = size_start + len(size_prefix)
    size_end = html.find(";\r\n", size_json_start)
    if size_end < 0:
        size_end = html.find(";\n", size_json_start)
    size_images = json.loads(html[size_json_start:size_end])
    return cards, size_images, card_start, size_end, card_prefix, size_prefix


def catalog_code_for(product: dict, size: str, size_images: dict) -> str | None:
    pid = product["id"]
    paths = [
        (size_images.get(pid) or {}).get(size),
        (product.get("vialImages") or {}).get(size),
        (product.get("catalogNos") or {}).get(size),
    ]
    for p in paths:
        if not p:
            continue
        if isinstance(p, str) and "/" not in p and not p.endswith(".png"):
            return p
        base = Path(str(p).replace("\\", "/")).name
        if base.endswith(".png"):
            return base[:-4]
        return base
    return None


def esc_size(size: str) -> str:
    return re.sub(r"[.*+?^${}()|[\]\\]", r"\\\g<0>", size)


def set_pill_price(card_html: str, size: str, price: str, code: str | None = None) -> str:
    if not card_html or not price:
        return card_html
    # Match the size-pill span for this size and set/replace data-price
    pattern = re.compile(
        rf'(<span class="size-pill[^"]*"[^>]*data-price=")[^"]*("([^>]*)>\s*{esc_size(size)}\s*</span>)'
    )
    if pattern.search(card_html):
        out = pattern.sub(rf"\1{price}\2", card_html, count=1)
    else:
        pattern2 = re.compile(
            rf'(<span class="size-pill[^"]*")([^>]*>\s*{esc_size(size)}\s*</span>)'
        )
        if pattern2.search(card_html):
            out = pattern2.sub(rf'\1 data-price="{price}"\2', card_html, count=1)
        else:
            return card_html
    if code:
        esc_code = re.sub(r"[.*+?^${}()|[\]\\]", r"\\\g<0>", code)
        pill_pat = re.compile(
            rf'(<span class="size-pill[^"]*")([^>]*>\s*{esc_size(size)}\s*</span>)'
        )
        if f'data-code="{code}"' in out:
            out = re.sub(rf'data-code="[^"]*"([^>]*>\s*{esc_size(size)}\s*</span>)', rf'data-code="{code}"\1', out, count=1)
        elif pill_pat.search(out):
            out = pill_pat.sub(rf'\1 data-code="{code}"\2', out, count=1)
    return out


def kit_price_for_codes(by_code: dict[str, dict], codes: str | list[str]) -> str | None:
    code_list = [codes] if isinstance(codes, str) else list(codes)
    prices = []
    for code in code_list:
        row = by_code.get(code)
        if row and row.get("kit"):
            prices.append(float(re.sub(r"[^0-9.]", "", row["kit"])))
    if not prices:
        return None
    lo = min(prices)
    if len(code_list) > 1 and len(prices) > 1 and lo != max(prices):
        return f"From {fmt_price(lo)}"
    return fmt_price(lo)


def sync_research_kit_prices(html: str, by_code: dict[str, dict]) -> tuple[str, list[str]]:
    changes: list[str] = []
    for title, codes in KIT_CARD_CODES.items():
        new_price = kit_price_for_codes(by_code, codes)
        if not new_price:
            continue
        esc_title = re.escape(title)
        pattern = re.compile(
            rf'(<h3>{esc_title}</h3>\s*<p>[^<]*</p>\s*<div class="kit-price">)[^<]*(</div>)'
        )
        match = pattern.search(html)
        if not match:
            continue
        old_price = match.group(0)
        old_val = re.search(r'kit-price">([^<]+)', old_price)
        old_text = old_val.group(1) if old_val else ""
        if old_text == new_price:
            continue
        html = pattern.sub(rf"\1{new_price}\2", html, count=1)
        changes.append(f"{title}: {old_text} -> {new_price}")
    return html, changes


def remove_pill(card_html: str, size: str) -> str:
    return re.sub(
        rf'<span class="size-pill[^"]*"[^>]*>\s*{esc_size(size)}\s*</span>\s*',
        "",
        card_html,
    )


def normalize_card_active(card_html: str) -> str:
    card_html = re.sub(r'class="size-pill active"', 'class="size-pill"', card_html)
    card_html = re.sub(
        r'class="size-pill([^"]*)\s+active"',
        r'class="size-pill\1"',
        card_html,
    )
    card_html = re.sub(r'(<span class="size-pill)([^"]*")', r'\1 active\2', card_html, count=1)
    return card_html


def prune_product_fields(product: dict, pid: str, keep_sizes: list[str], size_images: dict):
    product["vialSizes"] = keep_sizes[:]
    if product.get("prices"):
        product["prices"] = {k: v for k, v in product["prices"].items() if k in keep_sizes}
    if product.get("vialImages"):
        product["vialImages"] = {k: v for k, v in product["vialImages"].items() if k in keep_sizes}
    if product.get("catalogNos"):
        product["catalogNos"] = {k: v for k, v in product["catalogNos"].items() if k in keep_sizes}
    if pid in size_images:
        size_images[pid] = {k: v for k, v in size_images[pid].items() if k in keep_sizes}


def price_range(prices: dict[str, str]) -> str | None:
    vals = []
    for p in prices.values():
        n = float(re.sub(r"[^0-9.]", "", p))
        vals.append(n)
    if not vals:
        return None
    lo, hi = min(vals), max(vals)
    if lo == hi:
        return fmt_price(lo)
    return f"{fmt_price(lo)} – {fmt_price(hi)}"


def main():
    by_code = load_price_compare()
    products = json.loads(PRODUCTS.read_text(encoding="utf-8"))
    html = INDEX.read_text(encoding="utf-8")
    cards, size_images, card_start, size_end, card_prefix, size_prefix = parse_index_embeds(html)

    updated_products = 0
    updated_sizes = 0
    pruned_sizes = 0
    missing_codes = []
    changes = []

    us_products = [p for p in products if p.get("warehouse") == "US Warehouse"]
    for product in us_products:
        pid = product["id"]
        keep_sizes: list[str] = []
        prices: dict[str, str] = {}
        drop_sizes: list[str] = []

        for size in product.get("vialSizes") or []:
            code = catalog_code_for(product, size, size_images)
            if not code:
                missing_codes.append(f"{product['name']} {size}: no catalog code")
                drop_sizes.append(size)
                continue
            row = by_code.get(code)
            if not row or not row["single"]:
                missing_codes.append(f"{product['name']} {size}: code {code} not in PRICE_COMPARE")
                drop_sizes.append(size)
                continue
            keep_sizes.append(size)
            new_price = row["single"]
            old_price = (product.get("prices") or {}).get(size)
            prices[size] = new_price
            if old_price != new_price:
                changes.append(f"{product['name']} {size} ({code}): {old_price} -> {new_price}")
                updated_sizes += 1
            if pid in cards:
                cards[pid] = set_pill_price(cards[pid], size, new_price, code)
            product.setdefault("catalogNos", {})[size] = code

        if drop_sizes and pid in cards:
            for size in drop_sizes:
                cards[pid] = remove_pill(cards[pid], size)
                pruned_sizes += 1

        if keep_sizes:
            prune_product_fields(product, pid, keep_sizes, size_images)
            product["prices"] = prices
            rng = price_range(prices)
            if rng:
                product["price"] = rng
            if pid in cards:
                cards[pid] = normalize_card_active(cards[pid])
            updated_products += 1
        elif product.get("visible"):
            missing_codes.append(f"{product['name']}: no PRICE_COMPARE sizes kept")

    PRODUCTS.write_text(json.dumps(products, indent=2) + "\n", encoding="utf-8")

    line_term = size_prefix[: size_prefix.index("window")]
    new_block = (
        card_prefix
        + json.dumps(cards, ensure_ascii=False)
        + size_prefix
        + json.dumps(size_images, ensure_ascii=False)
        + line_term
    )
    updated_html = html[:card_start] + new_block + html[size_end + len(line_term) :]
    updated_html, kit_changes = sync_research_kit_prices(updated_html, by_code)
    INDEX.write_text(updated_html, encoding="utf-8")

    print(f"PRICE_COMPARE codes: {len(by_code)}")
    print(f"Products updated: {updated_products}")
    print(f"Size prices updated: {updated_sizes}")
    print("Sample changes:")
    for line in changes[:25]:
        print(" ", line)
    if len(changes) > 25:
        print(f"  ... +{len(changes) - 25} more")
    if kit_changes:
        print(f"Research kit prices updated: {len(kit_changes)}")
        for line in kit_changes:
            print(" ", line)
    if missing_codes:
        print(f"Missing/unmapped ({len(missing_codes)}):")
        for line in missing_codes[:20]:
            print(" ", line)
        if len(missing_codes) > 20:
            print(f"  ... +{len(missing_codes) - 20} more")

    # Spot-check
    for name in ["5-Amino-1MQ", "BPC-157"]:
        p = next((x for x in products if x.get("name") == name and x.get("visible")), None)
        if p:
            print(f"VERIFY {name}:", p.get("prices"))


if __name__ == "__main__":
    main()
